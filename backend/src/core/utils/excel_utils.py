from collections.abc import Callable
from datetime import UTC, datetime
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def export_to_excel[T](
    resource_name: str,
    field_map: dict[str, Callable[[T], Any]],
    items: list[T],
) -> bytes:
    """Build an Excel workbook from a list of items and return it as raw bytes.

    The sheet title is set to ``<resource_name> <YYYY-MM-DD>`` (today's UTC date).
    Column headers are derived from the keys of field_map (bold, auto-width up to 50
    characters). Each row is produced by calling the corresponding field_map value with
    the item.

    Args:
        resource_name: Human-readable name used as the sheet title prefix.
        field_map: Ordered mapping of column header → extractor callable.
        items: Records to export; one row per item.

    Returns:
        Raw bytes of the saved .xlsx workbook.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert isinstance(sheet, Worksheet)
    sheet.title = f"{resource_name} {datetime.now(UTC).strftime('%Y-%m-%d')}"

    headers = list(field_map.keys())
    sheet.append(headers)
    bold_font = Font(bold=True)
    for cell in sheet[1]:
        cell.font = bold_font

    for item in items:
        sheet.append([fn(item) for fn in field_map.values()])

    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        max_length = 0
        for cell in column_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except (AttributeError, TypeError):
                pass
        col_letter = get_column_letter(col_idx)
        sheet.column_dimensions[col_letter].width = min(max_length + 2, 50)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
