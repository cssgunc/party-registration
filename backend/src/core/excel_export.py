from io import BytesIO

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


class ExcelExporter:
    """Encapsulates openpyxl mechanics for generating Excel workbooks."""

    def __init__(self, sheet_title: str = "Sheet") -> None:
        self._workbook = openpyxl.Workbook()
        sheet = self._workbook.active
        assert isinstance(sheet, Worksheet)
        self._sheet: Worksheet = sheet
        self._sheet.title = sheet_title

    def set_headers(self, headers: list[str]) -> "ExcelExporter":
        """Append headers as the first row with bold font. Returns self for chaining."""
        self._sheet.append(headers)
        bold_font = Font(bold=True)
        for cell in self._sheet[1]:
            cell.font = bold_font
        return self

    def add_row(self, row: list) -> "ExcelExporter":
        """Append a data row. Returns self for chaining."""
        self._sheet.append(row)
        return self

    def to_bytes(self) -> bytes:
        """Auto-fit column widths (capped at 50 chars), save workbook, return bytes."""
        for col_idx, column_cells in enumerate(self._sheet.columns, start=1):
            max_length = 0
            for cell in column_cells:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except (AttributeError, TypeError):
                    pass
            col_letter = get_column_letter(col_idx)
            self._sheet.column_dimensions[col_letter].width = min(max_length + 2, 50)

        buffer = BytesIO()
        self._workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def format_phone(phone: str) -> str:
        """Format 10-digit numbers as (XXX) XXX-XXXX; pass through anything else unchanged."""
        digits = "".join(filter(str.isdigit, phone)) if phone else phone
        if digits and len(digits) == 10:
            return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
        return phone
